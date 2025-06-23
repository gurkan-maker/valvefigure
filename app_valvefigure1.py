import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image, ImageDraw
import base64
import tempfile
import os
from datetime import datetime
from fpdf import FPDF
from io import BytesIO
import openpyxl
import re

# Initialize session state
if 'valves' not in st.session_state:
    st.session_state.valves = []
if 'current_valve' not in st.session_state:
    st.session_state.current_valve = {}
if 'proposal_name' not in st.session_state:
    st.session_state.proposal_name = "Valve Proposal"
if 'proposal_date' not in st.session_state:
    st.session_state.proposal_date = datetime.now().strftime("%Y-%m-%d")
if 'client_name' not in st.session_state:
    st.session_state.client_name = ""
if 'proposal_items' not in st.session_state:
    st.session_state.proposal_items = []
    
# Load default logo
if 'logo_bytes' not in st.session_state:
    try:
        with open("logo.png", "rb") as f:
            st.session_state.logo_bytes = f.read()
    except FileNotFoundError:
        # If logo.png doesn't exist, create a placeholder
        img = Image.new('RGB', (300, 100), color=(73, 109, 137))
        draw = ImageDraw.Draw(img)
        draw.text((10, 10), "VASTAŞ Valve Solutions", fill=(255, 255, 255))
        img_bytes = BytesIO()
        img.save(img_bytes, format='PNG')
        st.session_state.logo_bytes = img_bytes.getvalue()

# ========================
# MATERIAL DATABASE
# ========================
MATERIAL_DB = {
    'Body/Bonnet': {
        'Carbon Steel': {'price': 150, 'color': '#A9A9A9'},
        'Stainless Steel 316': {'price': 300, 'color': '#C0C0C0'},
        'Duplex Steel': {'price': 450, 'color': '#4682B4'},
        'Super Duplex': {'price': 600, 'color': '#5F9EA0'},
        'Alloy 20': {'price': 700, 'color': '#6495ED'},
        'Hastelloy C': {'price': 1200, 'color': '#B0C4DE'},
    },
    'Ball': {
        'Stainless Steel 316': {'price': 200, 'color': '#C0C0C0'},
        'Stellite 6': {'price': 500, 'color': '#FF6347'},
        'Alloy 6': {'price': 800, 'color': '#FFA500'},
        'Hastelloy C': {'price': 1500, 'color': '#B0C4DE'},
    },
    'Stem': {
        'Stainless Steel 316': {'price': 100, 'color': '#C0C0C0'},
        '17-4 PH': {'price': 250, 'color': '#9370DB'},
        'Monel': {'price': 400, 'color': '#D8BFD8'},
        'Inconel 718': {'price': 600, 'color': '#DA70D6'},
    },
    'Seat': {
        'PTFE': {'price': 80, 'color': '#FFFFFF'},
        'RPTFE': {'price': 120, 'color': '#F0F8FF'},
        'PEEK': {'price': 200, 'color': '#ADD8E6'},
        'Metal': {'price': 300, 'color': '#D3D3D3'},
        'Graphoil': {'price': 350, 'color': '#708090'},
    },
    'Trim': {
        'Standard': {'price': 100, 'color': '#D3D3D3'},
        'Anti-cavitation': {'price': 300, 'color': '#87CEEB'},
        'Low Noise': {'price': 400, 'color': '#20B2AA'},
        'Cryogenic': {'price': 600, 'color': '#00BFFF'},
    }
}

# ========================
# EXCEL PRICE DATABASE
# ========================
def load_price_database():
    return {
        'Valve Size': {
            '0.5"': 500,
            '1"': 800,
            '1.5"': 1200,
            '2"': 1800,
            '3"': 2500,
            '4"': 3500,
            '6"': 5000,
            '8"': 7000,
            '10"': 10000,
            '12"': 14000,
        },
        'Pressure Rating': {
            '150#': 1.0,
            '300#': 1.2,
            '600#': 1.5,
            '900#': 1.8,
            '1500#': 2.2,
            '2500#': 3.0,
        },
        'Actuator Type': {
            'Pneumatic': 800,
            'Electric': 1500,
            'Hydraulic': 2500,
        },
        'Accessories': {
            'Positioner': 400,
            'Solenoid Valve': 200,
            'Limit Switches': 300,
            'Air Filter Regulator': 150,
            'Boosters': 350,
        }
    }

# ========================
# VALVE FIGURE GENERATION
# ========================
def generate_valve_figure(valve_data):
    # Create a blank image with white background
    img = Image.new('RGB', (400, 600), color='white')
    draw = ImageDraw.Draw(img)
    
    # Draw valve body
    body_color = MATERIAL_DB['Body/Bonnet'][valve_data['body_material']]['color']
    draw.rectangle([100, 100, 300, 400], fill=body_color, outline='black', width=2)
    
    # Draw ball
    ball_color = MATERIAL_DB['Ball'][valve_data['ball_material']]['color']
    draw.ellipse([150, 200, 250, 300], fill=ball_color, outline='black', width=2)
    
    # Draw stem
    stem_color = MATERIAL_DB['Stem'][valve_data['stem_material']]['color']
    draw.rectangle([190, 100, 210, 200], fill=stem_color, outline='black', width=1)
    
    # Draw actuator
    draw.rectangle([150, 50, 250, 100], fill='#FFD700', outline='black', width=2)
    
    # Add labels
    draw.text((200, 30), f"{valve_data['actuator_type']} Actuator", fill='black', anchor='ms')
    draw.text((200, 450), f"{valve_data['size']} {valve_data['pressure_rating']}", fill='black', anchor='ms')
    draw.text((200, 480), f"Body: {valve_data['body_material']}", fill='black', anchor='ms')
    draw.text((200, 510), f"Ball: {valve_data['ball_material']}", fill='black', anchor='ms')
    draw.text((200, 540), f"Stem: {valve_data['stem_material']}", fill='black', anchor='ms')
    
    return img

# ========================
# PRICE CALCULATION
# ========================
def calculate_valve_price(valve_data, price_db):
    try:
        # Base price based on size
        base_price = price_db['Valve Size'][valve_data['size']]
        
        # Pressure rating multiplier
        pressure_multiplier = price_db['Pressure Rating'][valve_data['pressure_rating']]
        
        # Material costs
        body_cost = MATERIAL_DB['Body/Bonnet'][valve_data['body_material']]['price']
        ball_cost = MATERIAL_DB['Ball'][valve_data['ball_material']]['price']
        stem_cost = MATERIAL_DB['Stem'][valve_data['stem_material']]['price']
        seat_cost = MATERIAL_DB['Seat'][valve_data['seat_material']]['price']
        trim_cost = MATERIAL_DB['Trim'][valve_data['trim_type']]['price']
        
        # Actuator cost
        actuator_cost = price_db['Actuator Type'][valve_data['actuator_type']]
        
        # Accessories cost
        accessories_cost = 0
        for accessory in valve_data['accessories']:
            accessories_cost += price_db['Accessories'].get(accessory, 0)
        
        # Calculate total price
        material_cost = body_cost + ball_cost + stem_cost + seat_cost + trim_cost
        total_price = (base_price + material_cost + actuator_cost + accessories_cost) * pressure_multiplier
        
        # Apply quantity
        total_price *= valve_data['quantity']
        
        return {
            'base_price': base_price,
            'pressure_multiplier': pressure_multiplier,
            'material_cost': material_cost,
            'actuator_cost': actuator_cost,
            'accessories_cost': accessories_cost,
            'total_price': total_price
        }
    except KeyError as e:
        st.error(f"Missing price data for: {str(e)}")
        return None

# ========================
# PDF REPORT GENERATION
# ========================
class PDF(FPDF):
    def header(self):
        if st.session_state.get('logo_bytes'):
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmpfile:
                    tmpfile.write(st.session_state.logo_bytes)
                    self.image(tmpfile.name, 10, 8, 25)
                os.unlink(tmpfile.name)
            except:
                pass
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, st.session_state.proposal_name, 0, 1, 'C')
        self.set_font('Arial', '', 12)
        self.cell(0, 10, f'Prepared for: {st.session_state.client_name}', 0, 1, 'C')
        self.cell(0, 10, f'Date: {st.session_state.proposal_date}', 0, 1, 'C')
        self.ln(10)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

def generate_proposal_pdf():
    pdf = PDF()
    pdf.add_page()
    
    # Add title
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Valve Configuration Proposal', 0, 1, 'C')
    pdf.ln(5)
    
    # Add client info
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Client: {st.session_state.client_name}', 0, 1)
    pdf.cell(0, 10, f'Proposal Date: {st.session_state.proposal_date}', 0, 1)
    pdf.cell(0, 10, f'Prepared By: VASTAŞ Valve Solutions', 0, 1)
    pdf.ln(10)
    
    # Add valve configurations
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Valve Configurations', 0, 1)
    pdf.ln(5)
    
    # Create table headers
    col_widths = [40, 30, 25, 25, 25, 25, 30]
    headers = ['Description', 'Size', 'Material', 'Actuator', 'Fail Mode', 'Qty', 'Price (€)']
    
    pdf.set_font('Arial', 'B', 10)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
    pdf.ln()
    
    # Add valve items
    pdf.set_font('Arial', '', 10)
    total_price = 0
    for item in st.session_state.proposal_items:
        description = f"{item['size']} {item['type']} Valve"
        material = item['body_material']
        pdf.cell(col_widths[0], 10, description, 1)
        pdf.cell(col_widths[1], 10, item['size'], 1)
        pdf.cell(col_widths[2], 10, material, 1)
        pdf.cell(col_widths[3], 10, item['actuator_type'], 1)
        pdf.cell(col_widths[4], 10, item['fail_mode'], 1)
        pdf.cell(col_widths[5], 10, str(item['quantity']), 1)
        pdf.cell(col_widths[6], 10, f"{item['total_price']:,.2f}", 1)
        pdf.ln()
        total_price += item['total_price']
    
    # Add total
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(sum(col_widths[:-1]), 10, 'TOTAL:', 1, 0, 'R')
    pdf.cell(col_widths[-1], 10, f"{total_price:,.2f}", 1, 1)
    
    # Add notes
    pdf.ln(10)
    pdf.set_font('Arial', 'I', 10)
    pdf.multi_cell(0, 5, 'Notes: Prices are in Euros (€). Delivery time is 8-12 weeks from order confirmation. Prices valid for 30 days.')
    
    # Save to bytes buffer
    pdf_bytes = BytesIO(pdf.output(dest='S').encode('latin1'))
    pdf_bytes.seek(0)
    return pdf_bytes

# ========================
# STREAMLIT UI
# ========================
def main():
    st.set_page_config(
        page_title="VASTAŞ Valve Configurator",
        page_icon="🔧",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS
    st.markdown("""
        <style>
        .stApp {
            background-color: #f0f2f6;
        }
        .block-container {
            padding-top: 1rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 10px;
        }
        .stTabs [data-baseweb="tab"] {
            height: 50px;
            padding: 15px 25px;
            border-radius: 10px 10px 0 0;
            font-size: 18px;
        }
        .stTabs [aria-selected="true"] {
            background-color: #1f77b4;
            color: white;
        }
        .stButton button {
            width: 100%;
            font-weight: bold;
            font-size: 18px;
        }
        .config-card {
            background-color: white;
            border-radius: 10px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .price-display {
            font-size: 24px;
            font-weight: bold;
            color: #1f77b4;
            text-align: center;
            padding: 15px;
            border: 2px solid #1f77b4;
            border-radius: 10px;
            margin: 20px 0;
        }
        .material-item {
            padding: 10px;
            margin: 5px 0;
            border-radius: 5px;
            border-left: 4px solid #1f77b4;
            background-color: #e6f0ff;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Load price database
    price_db = load_price_database()
    
    # Sidebar - Logo and Actions
    with st.sidebar:
        st.title("VASTAŞ Valve Configurator")
        
        # Display company logo
        st.image(st.session_state.logo_bytes, caption="Company Logo", use_column_width=True)
            
        st.divider()
        st.subheader("Actions")
        if st.button("Save Current Configuration"):
            if st.session_state.current_valve:
                st.session_state.valves.append(st.session_state.current_valve.copy())
                st.success("Valve configuration saved!")
            else:
                st.warning("No valve configuration to save")
                
        if st.button("Clear All Configurations"):
            st.session_state.valves = []
            st.session_state.proposal_items = []
            st.success("All configurations cleared")
            
        st.divider()
        st.subheader("Proposal Info")
        st.session_state.proposal_name = st.text_input("Proposal Name", st.session_state.proposal_name)
        st.session_state.client_name = st.text_input("Client Name", st.session_state.client_name)
        st.session_state.proposal_date = st.date_input("Proposal Date").strftime("%Y-%m-%d")
        
        if st.button("Generate Proposal PDF"):
            if st.session_state.proposal_items:
                pdf_bytes = generate_proposal_pdf()
                st.sidebar.download_button(
                    label="Download Proposal",
                    data=pdf_bytes,
                    file_name=f"{st.session_state.proposal_name.replace(' ', '_')}.pdf",
                    mime="application/pdf"
                )
            else:
                st.warning("No items in proposal")
    
    # Main Tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "⚙️ System Requirements", 
        "🛠️ Material Selection", 
        "💰 Pricing", 
        "📄 Proposal"
    ])
    
    # Tab 1: System Requirements
    with tab1:
        st.header("System Requirements")
        with st.form("system_requirements_form"):
            col1, col2 = st.columns(2)
            with col1:
                st.subheader("Operating Conditions")
                op_pressure = st.number_input("Operating Pressure (bar)", min_value=0.0, max_value=1000.0, value=50.0, step=1.0)
                op_temp = st.number_input("Operating Temperature (°C)", min_value=-200.0, max_value=1000.0, value=150.0, step=1.0)
                fluid_type = st.selectbox("Fluid Type", ["Water", "Steam", "Oil", "Gas", "Chemical", "Cryogenic"])
                
            with col2:
                st.subheader("Valve Specifications")
                valve_size = st.selectbox("Valve Size", list(price_db['Valve Size'].keys()))
                pressure_rating = st.selectbox("Pressure Rating", list(price_db['Pressure Rating'].keys()))
                valve_type = st.selectbox("Valve Type", ["Ball", "Globe", "Butterfly", "Gate", "Check"])
                fail_mode = st.selectbox("Fail Safe Mode", ["Fail to Close", "Fail to Open", "Fail in Position"])
                actuator_type = st.selectbox("Actuator Type", list(price_db['Actuator Type'].keys()))
            
            st.subheader("Additional Requirements")
            accessories = st.multiselect("Accessories", list(price_db['Accessories'].keys()))
            quantity = st.number_input("Quantity", min_value=1, max_value=100, value=1, step=1)
            notes = st.text_area("Special Requirements or Notes")
            
            if st.form_submit_button("Save Requirements"):
                st.session_state.current_valve.update({
                    'op_pressure': op_pressure,
                    'op_temp': op_temp,
                    'fluid_type': fluid_type,
                    'size': valve_size,
                    'pressure_rating': pressure_rating,
                    'type': valve_type,
                    'fail_mode': fail_mode,
                    'actuator_type': actuator_type,
                    'accessories': accessories,
                    'quantity': quantity,
                    'notes': notes
                })
                st.success("System requirements saved!")
    
    # Tab 2: Material Selection
    with tab2:
        st.header("Material Selection")
        
        if not st.session_state.current_valve:
            st.warning("Please complete System Requirements first")
            st.stop()
            
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.subheader("Material Options")
            body_material = st.selectbox("Body/Bonnet Material", list(MATERIAL_DB['Body/Bonnet'].keys()))
            ball_material = st.selectbox("Ball Material", list(MATERIAL_DB['Ball'].keys()))
            stem_material = st.selectbox("Stem Material", list(MATERIAL_DB['Stem'].keys()))
            seat_material = st.selectbox("Seat Material", list(MATERIAL_DB['Seat'].keys()))
            trim_type = st.selectbox("Trim Type", list(MATERIAL_DB['Trim'].keys()))
            
            if st.button("Save Materials"):
                st.session_state.current_valve.update({
                    'body_material': body_material,
                    'ball_material': ball_material,
                    'stem_material': stem_material,
                    'seat_material': seat_material,
                    'trim_type': trim_type
                })
                st.success("Material selection saved!")
                
            st.divider()
            st.subheader("Material Summary")
            st.markdown(f"<div class='material-item'>Body: {body_material}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='material-item'>Ball: {ball_material}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='material-item'>Stem: {stem_material}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='material-item'>Seat: {seat_material}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='material-item'>Trim: {trim_type}</div>", unsafe_allow_html=True)
        
        with col2:
            st.subheader("Valve Configuration")
            # Generate valve image
            valve_img = generate_valve_figure({
                'size': st.session_state.current_valve.get('size', '2"'),
                'pressure_rating': st.session_state.current_valve.get('pressure_rating', '150#'),
                'body_material': body_material,
                'ball_material': ball_material,
                'stem_material': stem_material,
                'actuator_type': st.session_state.current_valve.get('actuator_type', 'Pneumatic')
            })
            
            st.image(valve_img, caption="Valve Configuration", use_column_width=True)
            
            # Material colors key
            st.subheader("Material Color Key")
            col_col1, col_col2 = st.columns(2)
            for i, (material_type, materials) in enumerate(MATERIAL_DB.items()):
                cols = col_col1 if i % 2 == 0 else col_col2
                with cols:
                    st.markdown(f"**{material_type}**")
                    for mat, props in materials.items():
                        st.markdown(
                            f"<div style='background-color:{props['color']}; padding:5px; margin:2px; border-radius:5px;'>"
                            f"{mat} - €{props['price']}"
                            f"</div>", 
                            unsafe_allow_html=True
                        )
    
    # Tab 3: Pricing
    with tab3:
        st.header("Pricing")
        
        if not st.session_state.current_valve or 'body_material' not in st.session_state.current_valve:
            st.warning("Please complete Material Selection first")
            st.stop()
            
        # Calculate price
        price_data = calculate_valve_price(st.session_state.current_valve, price_db)
        
        if price_data:
            col1, col2, col3 = st.columns([1, 2, 1])
            
            with col2:
                st.markdown(f"<div class='price-display'>Total Price: €{price_data['total_price']:,.2f}</div>", unsafe_allow_html=True)
                
                # Price breakdown
                st.subheader("Price Breakdown")
                st.markdown(f"- **Base Valve Price**: €{price_data['base_price']:,.2f}")
                st.markdown(f"- **Pressure Multiplier ({st.session_state.current_valve['pressure_rating']})**: x{price_data['pressure_multiplier']:.2f}")
                st.markdown(f"- **Material Costs**: €{price_data['material_cost']:,.2f}")
                st.markdown(f"- **Actuator ({st.session_state.current_valve['actuator_type']})**: €{price_data['actuator_cost']:,.2f}")
                st.markdown(f"- **Accessories**: €{price_data['accessories_cost']:,.2f}")
                st.markdown(f"- **Quantity**: {st.session_state.current_valve['quantity']}")
                
                # Add to proposal button
                if st.button("Add to Proposal"):
                    valve_with_price = st.session_state.current_valve.copy()
                    valve_with_price['total_price'] = price_data['total_price']
                    st.session_state.proposal_items.append(valve_with_price)
                    st.success("Added to proposal!")
    
    # Tab 4: Proposal
    with tab4:
        st.header("Proposal Summary")
        
        if not st.session_state.proposal_items:
            st.info("No items in proposal yet. Add valves from the Pricing tab.")
            st.stop()
            
        # Display proposal items
        st.subheader("Valves in Proposal")
        for i, item in enumerate(st.session_state.proposal_items):
            with st.expander(f"Valve {i+1}: {item['size']} {item['type']} Valve - €{item['total_price']:,.2f}"):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Configuration**")
                    st.markdown(f"- Size: {item['size']}")
                    st.markdown(f"- Pressure Rating: {item['pressure_rating']}")
                    st.markdown(f"- Type: {item['type']}")
                    st.markdown(f"- Fail Mode: {item['fail_mode']}")
                    st.markdown(f"- Actuator: {item['actuator_type']}")
                    
                with col2:
                    st.markdown("**Materials**")
                    st.markdown(f"- Body: {item['body_material']}")
                    st.markdown(f"- Ball: {item['ball_material']}")
                    st.markdown(f"- Stem: {item['stem_material']}")
                    st.markdown(f"- Seat: {item['seat_material']}")
                    st.markdown(f"- Trim: {item['trim_type']}")
                
                st.markdown("**Pricing**")
                st.markdown(f"- Quantity: {item['quantity']}")
                st.markdown(f"- Total Price: €{item['total_price']:,.2f}")
                
                if st.button(f"Remove Valve {i+1}", key=f"remove_{i}"):
                    st.session_state.proposal_items.pop(i)
                    st.experimental_rerun()
        
        # Proposal total
        total_price = sum(item['total_price'] for item in st.session_state.proposal_items)
        st.subheader(f"Proposal Total: €{total_price:,.2f}")
        
        # Export buttons
        st.download_button(
            label="Export Proposal to Excel",
            data=pd.DataFrame(st.session_state.proposal_items).to_csv(index=False).encode('utf-8'),
            file_name=f"{st.session_state.proposal_name.replace(' ', '_')}.csv",
            mime="text/csv"
        )

if __name__ == "__main__":
    main()